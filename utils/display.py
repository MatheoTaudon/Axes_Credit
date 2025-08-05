import streamlit as st 
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def bouton_retour_accueil():
    """Affiche un bouton pour revenir à la page d’accueil"""
    st.button("⬅️ Retour à l'accueil", on_click=lambda: st.session_state.update(page="accueil"))


def bouton_export_excel(df: pd.DataFrame, nom_fichier: str = "export.xlsx", nom_feuille: str = "Axes"):
    buffer = BytesIO()
    wb = load_workbook("utils/template.xlsx")
    ws = wb.active

    start_row = 3 

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="📥 Télécharger Excel",
        data=buffer.getvalue(),
        file_name=nom_fichier,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def message_legal_axes():
    """Affiche l’avertissement réglementaire sur les axes"""
    st.markdown("""
    <div style='text-align:center; font-size:0.85em; color:gray;'>
        ⚠️ Les axes affichés sont fournis à titre indicatif uniquement. Ils ne constituent ni une offre ferme, ni une garantie d’exécution.
    </div>
    """, unsafe_allow_html=True)